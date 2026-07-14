#!/usr/bin/env python3
"""
One-off import: reads the PAZMUN 2026 registration Excel and loads
delegados/pajes/asesores + autoridades de comité into Supabase Postgres,
generating a random qr_code per person.

Run: python3 scripts/import_data.py
Requires DATABASE_URL in .env (direct Postgres connection, bypasses RLS).
"""
import os
import re
import secrets
import sys
import csv

import openpyxl
import psycopg2

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(
    os.path.expanduser("~"), "Downloads", "Inscripciones | PAZMUN 2026.xlsx"
)
OUT_CSV = os.path.join(ROOT, "inputs", "qr_export.csv")

VALID_AUTHORITY_ROLES = {
    "Presidencia",
    "Moderación",
    "Relatoría",
    "Subdirección | Cuerpo de Prensa",
    "Direccción | Cuerpo de Prensa",
}


def load_env():
    env_path = os.path.join(ROOT, ".env")
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k, v)


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    s = re.sub(r"\s+", " ", s)
    return s or None


def clean_yesno_field(flag, detail):
    """'No' -> None, 'Sí'/'Si' -> detail text (falls back to the flag itself)."""
    flag = clean(flag)
    if not flag or flag.lower() in ("no",):
        return None
    detail = clean(detail)
    return detail or flag


def clean_freeform_or_no(value):
    """Field is either 'No' or the actual free-text answer (e.g. allergy name)."""
    value = clean(value)
    if not value or value.lower() in ("no", "no ", "ninguna", "ninguno", "-"):
        return None
    return value


def make_qr_code():
    return secrets.token_urlsafe(8).replace("_", "").replace("-", "")[:10]


def parse_delegados(ws, rows_out, skipped):
    for r in range(2, ws.max_row + 1):
        tipo = clean(ws.cell(row=r, column=5).value)
        if not tipo:
            continue

        if tipo in ("Delegado/a | Corresponsal de Prensa", "Delegación"):
            name = " ".join(
                filter(None, [clean(ws.cell(row=r, column=7).value), clean(ws.cell(row=r, column=8).value)])
            )
            if not name:
                skipped.append((r, "delegado sin nombre"))
                continue
            rows_out.append(
                dict(
                    full_name=name,
                    role="delegado",
                    authority_role=None,
                    committee=clean(ws.cell(row=r, column=14).value),
                    institution=clean(ws.cell(row=r, column=9).value),
                    city=clean(ws.cell(row=r, column=11).value),
                    diet=clean_yesno_field(ws.cell(row=r, column=15).value, ws.cell(row=r, column=16).value),
                    allergy=clean_freeform_or_no(ws.cell(row=r, column=17).value),
                )
            )

        elif tipo == "Auxiliar de Comité (Paje)":
            name = " ".join(
                filter(None, [clean(ws.cell(row=r, column=43).value), clean(ws.cell(row=r, column=44).value)])
            )
            if not name:
                skipped.append((r, "paje sin nombre"))
                continue
            prefs = clean(ws.cell(row=r, column=53).value)
            committee = prefs.split(",")[0].strip() if prefs else None
            rows_out.append(
                dict(
                    full_name=name,
                    role="paje",
                    authority_role=None,
                    committee=committee,
                    institution=clean(ws.cell(row=r, column=45).value),
                    city=clean(ws.cell(row=r, column=46).value),
                    diet=clean_yesno_field(ws.cell(row=r, column=50).value, ws.cell(row=r, column=51).value),
                    allergy=clean_freeform_or_no(ws.cell(row=r, column=52).value),
                )
            )

        elif tipo == "Asesor/a":
            name = clean(ws.cell(row=r, column=18).value)
            if not name:
                skipped.append((r, "asesor sin nombre"))
                continue
            rows_out.append(
                dict(
                    full_name=name,
                    role="asesor",
                    authority_role=None,
                    committee=None,
                    institution=clean(ws.cell(row=r, column=20).value),
                    city=None,
                    diet=None,
                    allergy=None,
                )
            )
        else:
            skipped.append((r, f"tipo desconocido: {tipo}"))


def parse_autoridades(ws, rows_out, review_needed):
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(row=r, column=1).value)
        if not name:
            continue
        authority_role = clean(ws.cell(row=r, column=2).value)
        committee = clean(ws.cell(row=r, column=4).value)
        institution = clean(ws.cell(row=r, column=5).value)
        city = clean(ws.cell(row=r, column=6).value)

        rows_out.append(
            dict(
                full_name=name,
                role="autoridad",
                authority_role=authority_role,
                committee=committee,
                institution=institution,
                city=city,
                diet=None,
                allergy=None,
            )
        )

        # A handful of rows have a second person's name jammed into columns
        # H/I. Only trust it when column I is actually a known role — several
        # of these are free-text notes ("Jala más como mode...") instead of a
        # role, so those get flagged instead of guessed.
        second_name = clean(ws.cell(row=r, column=8).value)
        second_role = clean(ws.cell(row=r, column=9).value)
        if second_name:
            if second_role in VALID_AUTHORITY_ROLES:
                rows_out.append(
                    dict(
                        full_name=second_name,
                        role="autoridad",
                        authority_role=second_role,
                        committee=committee,
                        institution=institution,
                        city=city,
                        diet=None,
                        allergy=None,
                    )
                )
            else:
                review_needed.append((r, second_name, second_role))


def main():
    load_env()
    db_url = os.environ["DATABASE_URL"]

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    people = []
    skipped = []
    review_needed = []

    parse_delegados(wb["Delegados, pajes & asesores"], people, skipped)
    parse_autoridades(wb["Lista Oficial de Autoridades de"], people, review_needed)

    for p in people:
        p["qr_code"] = make_qr_code()

    conn = psycopg2.connect(db_url)
    cur = conn.cursor()
    cur.execute("delete from participants")  # idempotent re-import for this one-off script
    for p in people:
        cur.execute(
            """
            insert into participants
                (qr_code, full_name, role, authority_role, committee, institution, city, diet, allergy)
            values (%(qr_code)s, %(full_name)s, %(role)s, %(authority_role)s, %(committee)s,
                    %(institution)s, %(city)s, %(diet)s, %(allergy)s)
            """,
            p,
        )
    conn.commit()

    os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
    with open(OUT_CSV, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["full_name", "role", "authority_role", "committee", "qr_code"])
        for p in people:
            w.writerow([p["full_name"], p["role"], p["authority_role"] or "", p["committee"] or "", p["qr_code"]])

    from collections import Counter

    counts = Counter(p["role"] for p in people)
    print(f"Insertados: {len(people)} — {dict(counts)}")
    print(f"CSV con qr_code por persona: {OUT_CSV}")

    if skipped:
        print(f"\n{len(skipped)} filas omitidas (sin nombre o tipo desconocido):")
        for r, reason in skipped:
            print(f"  fila {r}: {reason}")

    if review_needed:
        print(f"\n{len(review_needed)} nombres secundarios en 'Lista Oficial de Autoridades' con rol ambiguo (NO importados, revisar a mano):")
        for r, name, role in review_needed:
            print(f"  fila {r}: {name!r} (texto en columna rol: {role!r})")

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
