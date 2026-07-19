#!/usr/bin/env python3
"""
Fills in `assignment` (country represented, or press outlet) from the
official Asignaciones sheet, matched by email. Mesa positions already have
this via authority_role so only 'País' and 'Medio de Comunicación' blocks
are used here.

Run: python3 scripts/add_assignments.py
"""
import os

import openpyxl
import psycopg2

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASIGNACIONES_PATH = os.path.join(
    os.path.expanduser("~"), "Downloads", "Asignaciones | PAZMUN 2026.xlsx"
)


def load_env():
    with open(os.path.join(ROOT, ".env")) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k, v)


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def assignment_by_email():
    wb = openpyxl.load_workbook(ASIGNACIONES_PATH, read_only=True, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        correo_idx = None
        header_kind = None
        last_label = None
        for row in rows:
            cells = [clean(c) for c in row]
            if not any(cells):
                last_label = None
                continue
            if "Correo electrónico" in cells:
                correo_idx = cells.index("Correo electrónico")
                first = cells[0]
                header_kind = (
                    "pais" if first == "País"
                    else "medio" if first == "Medio de Comunicación"
                    else "cargo"
                )
                last_label = None
                continue
            if correo_idx is None:
                continue
            label = cells[0]
            correo = cells[correo_idx]
            if label:
                last_label = label
            if header_kind in ("pais", "medio") and correo and last_label:
                result[correo.lower()] = last_label
    return result


def main():
    load_env()
    assignments = assignment_by_email()

    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    conn.autocommit = True
    cur = conn.cursor()
    cur.execute("select id, email from participants where email is not null")
    rows = cur.fetchall()

    updated = 0
    for pid, email in rows:
        value = assignments.get(email.lower())
        if value:
            cur.execute(
                "update participants set assignment = %s, updated_at = now() where id = %s",
                (value, pid),
            )
            updated += 1

    print(f"Asignaciones actualizadas: {updated} / {len(rows)}")
    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
