#!/usr/bin/env python3
"""
Full reconciliation of participants against the official Asignaciones
sheet: for anyone the sheet and our DB both know about (matched by email),
update committee/authority_role/assignment/full_name to the sheet's current
values. For anyone the sheet has that we don't, create a new participant
with a fresh QR code. People in our DB whose email the new sheet doesn't
mention at all are left untouched — we don't know what happened to them.

Run: python3 scripts/reconcile_asignaciones.py
"""
import os
import secrets

import openpyxl
import psycopg2

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASIGNACIONES_PATH = os.path.join(
    os.path.expanduser("~"), "Downloads", "Asignaciones | PAZMUN 2026.xlsx"
)

MESA_ROLE = {
    "Presidencia": ("autoridad", "Presidencia"),
    "Moderación": ("autoridad", "Moderación"),
    "Relatoría": ("autoridad", "Relatoría"),
    "Auxiliar de Comité": ("paje", None),
    "Dirección - Coordinación de Entrevistas": ("autoridad", "Dirección"),
    "Subdirección - Coordinación de Redes Sociales": ("autoridad", "Subdirección"),
    "Magistrado / Presidente": ("autoridad", "Magistrado / Presidente"),
    "Magistrado / Vicepresidente": ("autoridad", "Magistrado / Vicepresidente"),
}


def load_env():
    with open(os.path.join(ROOT, ".env")) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k, v)


def make_qr_code():
    return secrets.token_urlsafe(8).replace("_", "").replace("-", "")[:10]


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_asignaciones():
    """email -> dict(full_name, committee, role, authority_role, assignment)"""
    wb = openpyxl.load_workbook(ASIGNACIONES_PATH, read_only=True, data_only=True)
    people = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        committee_label = clean(rows[1][0]) if len(rows) > 1 else sheet_name
        correo_idx = None
        header_kind = None
        last_label = None
        for row in rows:
            cells = [clean(c) for c in row]
            if not any(cells):
                last_label = None
                continue
            if "Correo electrónico" in cells:
                correo_idx = cells.index("Correo electrónico")
                first = cells[0]
                header_kind = (
                    "pais" if first == "País"
                    else "medio" if first == "Medio de Comunicación"
                    else "cargo"
                )
                last_label = None
                continue
            if correo_idx is None:
                continue
            label, nombres, apellidos = cells[0], cells[1], cells[2]
            correo = cells[correo_idx]
            if not nombres and not apellidos:
                continue
            if label:
                last_label = label
            if not correo:
                continue
            full_name = f"{nombres or ''} {apellidos or ''}".strip()

            if header_kind == "cargo":
                role, authority_role = MESA_ROLE.get(last_label, ("autoridad", last_label))
                assignment = None
            elif header_kind == "pais":
                role, authority_role, assignment = "delegado", None, last_label
            else:  # medio
                role, authority_role, assignment = "delegado", None, last_label

            people[correo.lower()] = {
                "full_name": full_name,
                "committee": committee_label,
                "role": role,
                "authority_role": authority_role,
                "assignment": assignment,
            }
    return people


def main():
    load_env()
    sheet_people = parse_asignaciones()
    print(f"Personas parseadas de Asignaciones: {len(sheet_people)}")

    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    conn.autocommit = True
    cur = conn.cursor()
    cur.execute(
        "select id, full_name, email, role, committee, authority_role, assignment from participants"
    )
    db_rows = cur.fetchall()
    db_by_email = {r[2].lower(): r for r in db_rows if r[2]}

    updated = []
    created = []
    role_changes = []
    for email, sheet_p in sheet_people.items():
        if email in db_by_email:
            pid, db_name, _, db_role, db_committee, db_auth, db_assign = db_by_email[email]

            if sheet_p["role"] != db_role:
                # a role change (delegado <-> autoridad) is a bigger identity
                # shift than a field correction — leave the record untouched
                # and surface it for an explicit, deliberate update instead
                role_changes.append((pid, db_name, email, db_role, sheet_p))
                continue

            changes = {}
            if sheet_p["full_name"] and sheet_p["full_name"] != db_name:
                changes["full_name"] = sheet_p["full_name"]
            if sheet_p["committee"] != db_committee:
                changes["committee"] = sheet_p["committee"]
            if sheet_p["authority_role"] != db_auth:
                changes["authority_role"] = sheet_p["authority_role"]
            if sheet_p["assignment"] != db_assign:
                changes["assignment"] = sheet_p["assignment"]
            if changes:
                set_clause = ", ".join(f"{k} = %s" for k in changes)
                cur.execute(
                    f"update participants set {set_clause}, updated_at = now() where id = %s",
                    list(changes.values()) + [pid],
                )
                updated.append((db_name, email, changes))
        else:
            qr_code = make_qr_code()
            cur.execute(
                "insert into participants "
                "(qr_code, full_name, role, authority_role, committee, assignment, email) "
                "values (%s, %s, %s, %s, %s, %s, %s)",
                (
                    qr_code, sheet_p["full_name"], sheet_p["role"], sheet_p["authority_role"],
                    sheet_p["committee"], sheet_p["assignment"], email,
                ),
            )
            created.append((sheet_p["full_name"], email, sheet_p["role"], sheet_p["committee"]))

    # Role changes (delegado <-> autoridad): apply as one deliberate update
    # covering every field at once, since role alone isn't enough context.
    for pid, db_name, email, old_role, sheet_p in role_changes:
        cur.execute(
            "update participants set role = %s, committee = %s, authority_role = %s, "
            "assignment = %s, full_name = %s, updated_at = now() where id = %s",
            (
                sheet_p["role"], sheet_p["committee"], sheet_p["authority_role"],
                sheet_p["assignment"], sheet_p["full_name"] or db_name, pid,
            ),
        )

    # DB participants the new sheet never mentions at all
    unresolved = [
        r for r in db_rows if not r[2] or r[2].lower() not in sheet_people
    ]

    cur.close()
    conn.close()

    print(f"\nActualizados: {len(updated)}")
    for name, email, changes in updated:
        print(f"  {name} <{email}> -> {changes}")

    print(f"\nCambios de rol (delegado <-> autoridad): {len(role_changes)}")
    for pid, db_name, email, old_role, sheet_p in role_changes:
        print(f"  {db_name} <{email}>: era {old_role} -> ahora {sheet_p['role']} en {sheet_p['committee']} ({sheet_p['authority_role'] or sheet_p['assignment']})")

    print(f"\nCreados: {len(created)}")
    for name, email, role, committee in created:
        print(f"  {name} <{email}> — {role} — {committee}")

    print(f"\nSin corresponder en la hoja nueva (sin cambios): {len(unresolved)}")


if __name__ == "__main__":
    main()
