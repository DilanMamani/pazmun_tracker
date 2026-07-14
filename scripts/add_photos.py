#!/usr/bin/env python3
"""
Matches each existing participant against the registration Excel to pull
their photo Drive link, normalizes it to a direct-view URL, and UPDATEs
photo_url in place. Never touches qr_code or inserts/deletes rows —
credentials already generated stay valid.

Run: python3 scripts/add_photos.py
"""
import os
import re

import openpyxl
import psycopg2

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(os.path.expanduser("~"), "Downloads", "Inscripciones | PAZMUN 2026.xlsx")


def load_env():
    with open(os.path.join(ROOT, ".env")) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k, v)


def clean(v):
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v).strip())
    return s or None


def to_direct_view(url):
    url = clean(url)
    if not url:
        return None
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url) or re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    if not m:
        return None
    return f"https://drive.google.com/uc?export=view&id={m.group(1)}"


def normalize_name(name):
    return clean(name).lower() if name else ""


def collect_delegados_photos(ws):
    """(email, role, committee) -> photo_url, for delegado/paje/asesor rows."""
    out = {}
    for r in range(2, ws.max_row + 1):
        tipo = clean(ws.cell(row=r, column=5).value)
        if not tipo:
            continue
        email = clean(ws.cell(row=r, column=4).value)
        if not email:
            continue

        if tipo in ("Delegado/a | Corresponsal de Prensa", "Delegación"):
            role, committee, photo_col = "delegado", clean(ws.cell(row=r, column=14).value), 38
        elif tipo == "Auxiliar de Comité (Paje)":
            role, committee, photo_col = "paje", None, 38
        elif tipo == "Asesor/a":
            role, committee, photo_col = "asesor", None, 38
        else:
            continue

        photo = to_direct_view(ws.cell(row=r, column=photo_col).value)
        if photo:
            out[(email, role, committee)] = photo
            out.setdefault((email, role), photo)  # fallback key without committee
    return out


def collect_autoridades_photos(ws):
    """normalized full name -> photo_url, from the authority sign-up form."""
    out = {}
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(row=r, column=4).value)
        apellido = clean(ws.cell(row=r, column=5).value)
        full = " ".join(filter(None, [name, apellido]))
        photo = to_direct_view(ws.cell(row=r, column=10).value)
        if full and photo:
            out.setdefault(normalize_name(full), photo)
    return out


def main():
    load_env()
    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    cur = conn.cursor()
    cur.execute("select id, full_name, role, email, committee from participants")
    rows = cur.fetchall()

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    delegados_photos = collect_delegados_photos(wb["Delegados, pajes & asesores"])
    autoridad_photos = collect_autoridades_photos(wb["Autoridades de Comité"])

    updated, missing = 0, []

    for pid, full_name, role, email, committee in rows:
        photo = None
        if role == "autoridad":
            photo = autoridad_photos.get(normalize_name(full_name))
        elif email:
            photo = delegados_photos.get((email, role, committee)) or delegados_photos.get((email, role))

        if photo:
            cur.execute("update participants set photo_url = %s where id = %s", (photo, pid))
            updated += 1
        else:
            missing.append((full_name, role))

    conn.commit()
    print(f"Actualizados con foto: {updated} / {len(rows)}")
    if missing:
        print(f"\n{len(missing)} sin foto encontrada:")
        for name, role in missing:
            print(f"  - {name} ({role})")

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
