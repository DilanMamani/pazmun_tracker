#!/usr/bin/env python3
"""
Generates one PNG QR code per participant, pointing to their public profile
page. Files are named after the person's email, sorted into one folder per
committee (sourced from the official Asignaciones sheet — the DB's own
`committee` column has inconsistent naming across import batches), with
everyone the sheet doesn't cover landing in "Otros". Everything is zipped
into a single file at the end.

Run: python3 scripts/generate_qrcodes.py
Requires DATABASE_URL in .env.
"""
import os
import re
import shutil
import zipfile

import openpyxl
import psycopg2
import qrcode

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_URL = "https://pazmun-tracker.vercel.app"
OUT_DIR = os.path.join(ROOT, "inputs", "qr_codes")
ZIP_PATH = os.path.join(ROOT, "inputs", "qr_codes.zip")
ASIGNACIONES_PATH = os.path.join(
    os.path.expanduser("~"), "Downloads", "Asignaciones | PAZMUN 2026.xlsx"
)
OTROS = "Otros"


def load_env():
    env_path = os.path.join(ROOT, ".env")
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k, v)


def safe_filename(value):
    return re.sub(r"[^A-Za-z0-9@._-]", "_", value)


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def committee_by_email():
    """Email -> official committee name, parsed from the Asignaciones sheet.
    Each sheet mixes two header styles ('Cargo' for the mesa, 'País' or
    'Medio de Comunicación' for delegates/press) and some rows pair up under
    a shared label (e.g. two press people per outlet), so the label column
    is carried forward until a blank row resets it.
    """
    wb = openpyxl.load_workbook(ASIGNACIONES_PATH, read_only=True, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        committee_label = clean(rows[1][0]) if len(rows) > 1 else sheet_name
        correo_idx = None
        for row in rows:
            cells = [clean(c) for c in row]
            if not any(cells):
                continue
            if "Correo electrónico" in cells:
                correo_idx = cells.index("Correo electrónico")
                continue
            if correo_idx is None:
                continue
            correo = cells[correo_idx]
            if correo:
                result[correo.lower()] = committee_label
    return result


def main():
    load_env()
    db_url = os.environ["DATABASE_URL"]

    committee_map = committee_by_email()

    if os.path.exists(OUT_DIR):
        shutil.rmtree(OUT_DIR)
    os.makedirs(OUT_DIR)

    conn = psycopg2.connect(db_url)
    cur = conn.cursor()
    cur.execute("select full_name, role, email, qr_code from participants order by role, full_name")
    rows = cur.fetchall()
    cur.close()
    conn.close()

    used_names = {}
    used_name_fallback = []
    by_folder = {}
    generated = 0

    for full_name, role, email, qr_code in rows:
        if email:
            base = safe_filename(email)
            folder_label = committee_map.get(email.lower(), OTROS)
        else:
            base = safe_filename(full_name)
            folder_label = OTROS
            used_name_fallback.append(full_name)

        folder_name = safe_filename(folder_label)[:80]
        folder_path = os.path.join(OUT_DIR, folder_name)
        os.makedirs(folder_path, exist_ok=True)
        by_folder.setdefault(folder_label, 0)
        by_folder[folder_label] += 1

        count = used_names.get(base, 0)
        used_names[base] = count + 1
        filename = f"{base}.png" if count == 0 else f"{base}-{count + 1}.png"

        url = f"{BASE_URL}/p/{qr_code}"
        img = qrcode.make(url, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=12, border=3)
        img.save(os.path.join(folder_path, filename))
        generated += 1

    if os.path.exists(ZIP_PATH):
        os.remove(ZIP_PATH)
    with zipfile.ZipFile(ZIP_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        for dirpath, _, filenames in os.walk(OUT_DIR):
            for fname in filenames:
                full_path = os.path.join(dirpath, fname)
                arcname = os.path.relpath(full_path, OUT_DIR)
                zf.write(full_path, arcname)

    print(f"Generados {generated} PNG en {len(by_folder)} carpetas dentro de {OUT_DIR}")
    print(f"Zip: {ZIP_PATH}\n")
    for label, count in sorted(by_folder.items(), key=lambda x: x[0]):
        print(f"  {label}: {count}")

    dupes = {k: v for k, v in used_names.items() if v > 1}
    if dupes:
        print(f"\n{len(dupes)} nombres de archivo repetidos (sufijo -2, revisar si son la misma persona):")
        for base, count in dupes.items():
            print(f"  - {base} ({count} personas)")

    if used_name_fallback:
        print(f"\n{len(used_name_fallback)} personas sin email — se usó su nombre como archivo, van a 'Otros':")
        for name in used_name_fallback:
            print(f"  - {name}")


if __name__ == "__main__":
    main()
