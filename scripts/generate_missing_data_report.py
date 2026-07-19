#!/usr/bin/env python3
"""
Lists who has no photo and who never answered the allergy question at all
(blank field — distinct from someone who explicitly answered "ninguna"),
for follow-up before the event.

Run: python3 scripts/generate_missing_data_report.py
"""
import os

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_PATH = os.path.join(ROOT, "inputs", "PAZMUN_2026_fotos_y_alergias_faltantes.xlsx")


def load_env():
    with open(os.path.join(ROOT, ".env")) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k, v)


def role_label(role, authority_role):
    if role == "autoridad" and authority_role:
        return authority_role
    return {"delegado": "Delegado/a", "paje": "Paje", "asesor": "Asesor/a", "autoridad": "Autoridad"}[role]


def write_sheet(ws, rows, headers, widths):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A2D49")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for row in rows:
        ws.append(row)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    load_env()
    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    cur = conn.cursor()

    cur.execute(
        "select full_name, role, authority_role, committee, email from participants "
        "where photo_url is null order by role, committee, full_name"
    )
    no_photo = [
        [name, role_label(role, auth), committee or "", email or ""]
        for name, role, auth, committee, email in cur.fetchall()
    ]

    cur.execute(
        "select full_name, role, authority_role, committee, email from participants "
        "where allergy is null or allergy = '' order by role, committee, full_name"
    )
    no_allergy = [
        [name, role_label(role, auth), committee or "", email or ""]
        for name, role, auth, committee, email in cur.fetchall()
    ]

    cur.close()
    conn.close()

    wb = Workbook()
    write_sheet(
        wb.active, no_photo,
        ["Nombre completo", "Rol", "Comité", "Correo"], [28, 16, 45, 30],
    )
    wb.active.title = "Sin foto"

    write_sheet(
        wb.create_sheet("Sin alergia registrada"), no_allergy,
        ["Nombre completo", "Rol", "Comité", "Correo"], [28, 16, 45, 30],
    )

    wb.save(OUT_PATH)
    print(f"Generado: {OUT_PATH}")
    print(f"  Sin foto: {len(no_photo)}")
    print(f"  Sin alergia registrada (campo en blanco): {len(no_allergy)}")


if __name__ == "__main__":
    main()
