#!/usr/bin/env python3
"""
Builds the handoff spreadsheet for whoever produces the physical
credentials: full name, role, committee, what they represent, email, QR
code and credential link — split by role group.

Run: python3 scripts/generate_handoff_xlsx.py
"""
import os

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_URL = "https://pazmun-tracker.vercel.app"
OUT_PATH = os.path.join(ROOT, "inputs", "PAZMUN_2026_links_credenciales.xlsx")

HEADERS = [
    "Nombre completo", "Rol", "Comité", "Representa", "Institución",
    "Email", "Código QR", "Link de la credencial",
]


def load_env():
    with open(os.path.join(ROOT, ".env")) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k, v)


def role_label(role, authority_role):
    if role == "autoridad" and authority_role:
        return authority_role
    return {"delegado": "Delegado/a", "paje": "Paje", "asesor": "Asesor/a", "autoridad": "Autoridad"}[role]


def write_sheet(ws, rows):
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A2D49")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    for full_name, role, authority_role, committee, assignment, institution, email, qr_code in rows:
        ws.append([
            full_name,
            role_label(role, authority_role),
            committee or "",
            assignment or "",
            institution or "",
            email or "",
            qr_code,
            f"{BASE_URL}/p/{qr_code}",
        ])

    widths = [28, 16, 45, 30, 30, 30, 14, 42]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    load_env()
    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    cur = conn.cursor()
    cur.execute(
        "select full_name, role, authority_role, committee, assignment, institution, email, qr_code "
        "from participants order by role, committee, full_name"
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    autoridades = [r for r in rows if r[1] == "autoridad"]
    resto = [r for r in rows if r[1] != "autoridad"]

    wb = Workbook()
    write_sheet(wb.active, autoridades)
    wb.active.title = "Autoridades de Comité"
    write_sheet(wb.create_sheet("Delegados, Pajes y Asesores"), resto)

    wb.save(OUT_PATH)
    print(f"Generado: {OUT_PATH}")
    print(f"  Autoridades de Comité: {len(autoridades)}")
    print(f"  Delegados, Pajes y Asesores: {len(resto)}")


if __name__ == "__main__":
    main()
