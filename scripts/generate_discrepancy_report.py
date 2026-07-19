#!/usr/bin/env python3
"""
After scripts/reconcile_asignaciones.py runs, this lists whoever is STILL
unresolved: participants whose email the Asignaciones sheet never mentions
at all, with a reason for why they're in this state.

Run: python3 scripts/generate_discrepancy_report.py
"""
import os

import openpyxl
import psycopg2
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASIGNACIONES_PATH = os.path.join(
    os.path.expanduser("~"), "Downloads", "Asignaciones | PAZMUN 2026.xlsx"
)
OUT_PATH = os.path.join(ROOT, "inputs", "PAZMUN_2026_casos_a_revisar.xlsx")

# Old, short committee labels our autoridad-without-email records still
# carry (never reconciled, since there's no email to match on) -> who the
# sheet says replaced them, for a concrete, useful reason.
REPLACED_BY = {
    "Consejo Europeo": "Bruno Farias Sanchez (Presidencia, CEPE/UNECE)",
    "CSW (Mujeres)": "María Concepción Franco Rojas (Presidencia, CSW)",
    "SOCHUM": "Sara Edda Viscarra Sejas (Presidencia, SOCHUM)",
    "Ciencia y tecnología": "Marcos Yamil Salas Vedia (Presidencia, CSTD)",
}


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def load_env():
    with open(os.path.join(ROOT, ".env")) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k, v)


def sheet_emails():
    wb = openpyxl.load_workbook(ASIGNACIONES_PATH, read_only=True, data_only=True)
    emails = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        correo_idx = None
        for row in ws.iter_rows(values_only=True):
            cells = [clean(c) for c in row]
            if not any(cells):
                continue
            if "Correo electrónico" in cells:
                correo_idx = cells.index("Correo electrónico")
                continue
            if correo_idx is None:
                continue
            correo = cells[correo_idx]
            if correo:
                emails.add(correo.lower())
    return emails


SPECIAL_CASES = {
    "molinatenorioj@gmail.com": (
        "La hoja de Asignaciones tiene su nombre en 'Auxiliar de Comité' de UNICEF, pero con "
        "el correo de otra persona (Lucía Micaela Martinez Arévalo) pegado por error — "
        "actualizamos su comité a UNICEF por la coincidencia de nombre, pero su correo nunca "
        "quedó confirmado por la hoja. Verificar directamente con ella."
    ),
}


def reason(role, committee, authority_role, email):
    if email and email in SPECIAL_CASES:
        return SPECIAL_CASES[email]
    if role == "asesor":
        return "Asesor/a — no participa en un comité, la hoja de Asignaciones no cubre este rol."
    if not email:
        replacement = REPLACED_BY.get(committee)
        if replacement:
            return f"Reemplazado/a en la mesa directiva por {replacement}. No tenía correo registrado para confirmar si sigue participando."
        return "Sin correo registrado en nuestra base — no se pudo cruzar contra la hoja de Asignaciones."
    return "No aparece en la hoja de Asignaciones — verificar con el Secretariado si sigue participando y en qué comité."


def main():
    load_env()
    emails_in_sheet = sheet_emails()

    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    cur = conn.cursor()
    cur.execute(
        "select full_name, role, committee, authority_role, email from participants "
        "order by role, committee, full_name"
    )
    rows = [r for r in cur.fetchall() if not r[4] or r[4].lower() not in emails_in_sheet]
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Casos sin resolver"
    ws.append(["Nombre completo", "Rol", "Comité (en nuestra base)", "Cargo", "Correo", "Motivo"])
    for full_name, role, committee, authority_role, email in rows:
        ws.append([
            full_name, role, committee or "", authority_role or "", email or "",
            reason(role, committee, authority_role, email),
        ])

    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A2D49")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(6)}1"
    for i, w in enumerate([28, 12, 32, 16, 30, 60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT_PATH)
    print(f"Generado: {OUT_PATH}")
    print(f"  Casos sin resolver: {len(rows)}")


if __name__ == "__main__":
    main()
